"""Phase 11 — dashboard KPIs, forecast, reports, /ask."""
from collections import Counter
from datetime import date, datetime, timedelta
from decimal import Decimal, ROUND_HALF_UP
from typing import Literal, Optional

from fastapi import APIRouter, Depends, HTTPException, Query, Response
from pydantic import BaseModel
from sqlalchemy import and_, func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.deps import get_current_user, get_session, get_tenant, require_role
from app.models.menu_item import MenuItem
from app.models.stock_movement import StockMovement
from app.models.transaction import Transaction, TransactionItem
from app.models.user import User
from app.services.forecast import (daily_series_from_rows, eoq, ewma,
                                     reorder_point, safety_stock, z_score_anomaly)
from app.services.chenki_assistant import ask_owner

router = APIRouter(tags=["dashboard"])


def _range_for(range_key: Optional[str], frm: Optional[date], to: Optional[date]) -> tuple[datetime, datetime]:
    today = date.today()
    if range_key == "7d":
        start = today - timedelta(days=6)
    elif range_key == "30d":
        start = today - timedelta(days=29)
    elif range_key == "today" or range_key is None and frm is None:
        start = today
    elif frm is not None:
        start = frm
    else:
        start = today
    end = to or today
    return datetime.combine(start, datetime.min.time()), datetime.combine(end, datetime.max.time())


class DashboardOut(BaseModel):
    range: str
    revenue: Decimal
    tx_count: int
    avg_basket: Decimal
    top_item: Optional[str] = None
    low_stock: list[dict]
    staff_performance: list[dict]
    recent_transactions: list[dict]
    source_breakdown: dict[str, int]
    anomaly_z: float


@router.get("/dashboard", dependencies=[Depends(require_role("owner"))])
async def dashboard(
    range: Literal["today", "7d", "30d", "custom"] = "today",
    frm: Optional[date] = Query(default=None, alias="from"),
    to: Optional[date] = Query(default=None),
    tenant_id: int = Depends(get_tenant),
    session: AsyncSession = Depends(get_session),
) -> DashboardOut:
    start, end = _range_for(range, frm, to)

    tx_rows = (await session.execute(
        select(Transaction).where(
            Transaction.status == "paid",
            Transaction.paid_at >= start, Transaction.paid_at <= end,
        ).order_by(Transaction.paid_at.desc())
    )).scalars().all()

    revenue = sum((t.total for t in tx_rows), Decimal("0"))
    tx_count = len(tx_rows)
    avg_basket = ((revenue / tx_count) if tx_count else Decimal("0")).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)

    # Top item + source breakdown
    top_counter: Counter[int] = Counter()
    source_counter: Counter[str] = Counter()
    for tx in tx_rows:
        items = (await session.execute(select(TransactionItem).where(TransactionItem.transaction_id == tx.id))).scalars().all()
        for ti in items:
            if ti.menu_item_id:
                top_counter[ti.menu_item_id] += ti.quantity or 0
            if ti.source:
                source_counter[ti.source] += 1

    top_name = None
    if top_counter:
        best_id, _ = top_counter.most_common(1)[0]
        m = (await session.execute(select(MenuItem).where(MenuItem.id == best_id))).scalars().first()
        top_name = m.name if m else None

    low_stock_rows = (await session.execute(
        select(MenuItem).where(MenuItem.is_active.is_(True)).order_by(MenuItem.name)
    )).scalars().all()
    low_stock = [
        {"id": m.id, "name": m.name, "stock_qty": m.stock_qty or 0,
         "reorder_point": m.reorder_point or 0}
        for m in low_stock_rows if (m.stock_qty or 0) <= (m.reorder_point or 0)
    ]

    # Staff performance (count + revenue by staff_id)
    staff_counter: dict[int, tuple[int, Decimal]] = {}
    for tx in tx_rows:
        if tx.staff_id:
            c, r = staff_counter.get(tx.staff_id, (0, Decimal("0")))
            staff_counter[tx.staff_id] = (c + 1, r + tx.total)
    staff_performance: list[dict] = []
    for sid, (c, r) in staff_counter.items():
        u = (await session.execute(select(User).where(User.id == sid))).scalars().first()
        staff_performance.append({
            "staff_id": sid,
            "username": u.username if u else "?",
            "tx_count": c,
            "revenue": r,
        })

    recent = [
        {"id": tx.id, "tx_number": tx.tx_number, "total": tx.total,
         "staff_id": tx.staff_id, "paid_at": tx.paid_at}
        for tx in tx_rows[:10]
    ]

    # Anomaly (z-score of today revenue vs trailing 30 days)
    today = date.today()
    rows_30 = (await session.execute(
        select(Transaction.paid_at, Transaction.total).where(
            Transaction.status == "paid",
            Transaction.paid_at >= datetime.combine(today - timedelta(days=29), datetime.min.time()),
            Transaction.paid_at <= datetime.combine(today, datetime.max.time()),
        )
    )).all()
    window = daily_series_from_rows([(r[0].date(), float(r[1])) for r in rows_30], 30)
    z = z_score_anomaly(window[-1], window[:-1])

    return DashboardOut(
        range=range, revenue=revenue, tx_count=tx_count, avg_basket=avg_basket,
        top_item=top_name, low_stock=low_stock,
        staff_performance=staff_performance, recent_transactions=recent,
        source_breakdown=dict(source_counter), anomaly_z=round(z, 2),
    )


@router.get("/dashboard/charts", dependencies=[Depends(require_role("owner"))])
async def dashboard_charts(
    range: Literal["today", "7d", "30d", "custom"] = "7d",
    frm: Optional[date] = Query(default=None, alias="from"),
    to: Optional[date] = Query(default=None),
    tenant_id: int = Depends(get_tenant),
    session: AsyncSession = Depends(get_session),
) -> dict:
    """Chart data for the analytical dashboard:
    - item_sales: [{name, qty, revenue}] sorted desc by revenue (for pie chart)
    - daily_sales: [{date, revenue}] one row per day in range (for line chart)
    """
    start, end = _range_for(range, frm, to)

    ti_rows = (await session.execute(
        select(
            TransactionItem.menu_item_id,
            TransactionItem.quantity,
            TransactionItem.unit_price,
        ).join(Transaction, Transaction.id == TransactionItem.transaction_id).where(
            Transaction.status == "paid",
            Transaction.paid_at >= start,
            Transaction.paid_at <= end,
        )
    )).all()

    item_qty: Counter[int] = Counter()
    item_rev: dict[int, Decimal] = {}
    for mid, qty, price in ti_rows:
        if not mid:
            continue
        q = qty or 0
        item_qty[mid] += q
        item_rev[mid] = item_rev.get(mid, Decimal("0")) + (price or Decimal("0")) * q

    item_names = {
        m.id: m.name for m in
        (await session.execute(select(MenuItem))).scalars().all()
    }
    item_sales = sorted(
        [
            {"name": item_names.get(mid, f"#{mid}"),
             "qty": item_qty[mid],
             "revenue": float(item_rev.get(mid, Decimal("0")))}
            for mid in item_qty
        ],
        key=lambda r: r["revenue"], reverse=True,
    )

    tx_rows = (await session.execute(
        select(Transaction.paid_at, Transaction.total).where(
            Transaction.status == "paid",
            Transaction.paid_at >= start,
            Transaction.paid_at <= end,
        )
    )).all()

    # Seed every day in range with 0 so the line chart has no gaps
    by_day: dict[date, Decimal] = {}
    d = start.date()
    while d <= end.date():
        by_day[d] = Decimal("0")
        d += timedelta(days=1)
    for paid_at, total in tx_rows:
        if paid_at:
            by_day[paid_at.date()] = by_day.get(paid_at.date(), Decimal("0")) + (total or Decimal("0"))

    daily_sales = [
        {"date": d.isoformat(), "revenue": float(v)}
        for d, v in sorted(by_day.items())
    ]

    return {
        "range": range,
        "item_sales": item_sales,
        "daily_sales": daily_sales,
    }


@router.get("/dashboard/mobile", dependencies=[Depends(require_role("owner"))])
async def dashboard_mobile(tenant_id: int = Depends(get_tenant),
                             session: AsyncSession = Depends(get_session)) -> dict:
    """Slim version for the mobile owner light dashboard."""
    start, end = _range_for("today", None, None)
    tx_rows = (await session.execute(
        select(Transaction).where(
            Transaction.status == "paid",
            Transaction.paid_at >= start, Transaction.paid_at <= end,
        )
    )).scalars().all()
    revenue = sum((t.total for t in tx_rows), Decimal("0"))
    low_stock_count = (await session.execute(
        select(func.count()).select_from(MenuItem).where(
            MenuItem.is_active.is_(True),
            MenuItem.stock_qty <= MenuItem.reorder_point,
        )
    )).scalar() or 0
    return {"today_revenue": revenue, "today_tx": len(tx_rows),
            "low_stock_count": int(low_stock_count)}


@router.get("/forecast", dependencies=[Depends(require_role("owner"))])
async def forecast(tenant_id: int = Depends(get_tenant),
                    session: AsyncSession = Depends(get_session)) -> list[dict]:
    # 30-day daily demand per item, EWMA + SS + ROP + EOQ.
    items = (await session.execute(select(MenuItem).where(MenuItem.is_active.is_(True)))).scalars().all()
    since = datetime.combine(date.today() - timedelta(days=29), datetime.min.time())
    tx_items = (await session.execute(
        select(TransactionItem.menu_item_id, Transaction.paid_at, TransactionItem.quantity).join(
            Transaction, Transaction.id == TransactionItem.transaction_id
        ).where(Transaction.status == "paid", Transaction.paid_at >= since)
    )).all()

    by_item: dict[int, list[tuple[date, float]]] = {}
    for mid, pa, q in tx_items:
        if mid and pa:
            by_item.setdefault(mid, []).append((pa.date(), float(q or 0)))

    out: list[dict] = []
    for m in items:
        series = daily_series_from_rows(by_item.get(m.id, []), 30)
        e = ewma(series, 0.3)
        ss = safety_stock(series)
        rp = reorder_point(e, lead_time_days=7, ss=ss)
        annual = e * 365
        q = eoq(annual)
        out.append({
            "menu_item_id": m.id, "name": m.name,
            "ewma_daily": round(e, 2),
            "safety_stock": round(ss, 2),
            "reorder_point": round(rp, 1),
            "current_stock": m.stock_qty or 0,
            "eoq": q,
        })
    return out


@router.get("/reports", dependencies=[Depends(require_role("owner"))])
async def reports(
    format: Literal["csv", "json", "xlsx", "pdf"] = "csv",
    tenant_id: int = Depends(get_tenant),
    session: AsyncSession = Depends(get_session),
) -> Response:
    """Export 4 sections: sales-by-day, item performance, staff performance,
    inventory valuation. Formats: csv, json, xlsx, pdf."""
    since = datetime.combine(date.today() - timedelta(days=29), datetime.min.time())
    tx = (await session.execute(
        select(Transaction).where(Transaction.paid_at >= since, Transaction.status == "paid")
    )).scalars().all()

    # Section 1: sales by day
    by_day: dict[str, Decimal] = {}
    for t in tx:
        d = t.paid_at.strftime("%Y-%m-%d") if t.paid_at else "?"
        by_day[d] = by_day.get(d, Decimal("0")) + t.total

    # Section 2: item performance (qty + revenue per menu_item)
    item_qty: Counter[int] = Counter()
    item_rev: dict[int, Decimal] = {}
    for t in tx:
        tis = (await session.execute(
            select(TransactionItem).where(TransactionItem.transaction_id == t.id)
        )).scalars().all()
        for ti in tis:
            if not ti.menu_item_id:
                continue
            item_qty[ti.menu_item_id] += ti.quantity or 0
            line = (ti.unit_price or Decimal("0")) * (ti.quantity or 0) if hasattr(ti, "unit_price") else Decimal("0")
            item_rev[ti.menu_item_id] = item_rev.get(ti.menu_item_id, Decimal("0")) + line
    item_names = {m.id: m.name for m in (await session.execute(select(MenuItem))).scalars().all()}
    item_perf = [
        {"name": item_names.get(mid, f"#{mid}"), "qty": q, "revenue": str(item_rev.get(mid, Decimal("0")))}
        for mid, q in item_qty.most_common()
    ]

    # Section 3: staff performance
    staff_stats: dict[int, tuple[int, Decimal]] = {}
    for t in tx:
        if t.staff_id:
            c, r = staff_stats.get(t.staff_id, (0, Decimal("0")))
            staff_stats[t.staff_id] = (c + 1, r + t.total)
    staff_perf = []
    for sid, (c, r) in staff_stats.items():
        u = (await session.execute(select(User).where(User.id == sid))).scalars().first()
        staff_perf.append({"username": u.username if u else f"#{sid}", "tx_count": c, "revenue": str(r)})

    # Section 4: inventory valuation
    items = (await session.execute(select(MenuItem).where(MenuItem.is_active.is_(True)))).scalars().all()
    valuation_rows = [
        {"name": m.name, "stock_qty": m.stock_qty or 0,
         "avg_cost": str(m.avg_cost or Decimal("0")),
         "value": str((m.stock_qty or 0) * (m.avg_cost or Decimal("0")))}
        for m in items
    ]
    valuation_total = sum(((m.stock_qty or 0) * (m.avg_cost or Decimal("0")) for m in items), Decimal("0"))

    if format == "json":
        import json
        return Response(
            content=json.dumps({
                "sales_by_day": {k: str(v) for k, v in by_day.items()},
                "item_performance": item_perf,
                "staff_performance": staff_perf,
                "inventory_valuation_rows": valuation_rows,
                "inventory_valuation_total": str(valuation_total),
            }),
            media_type="application/json",
        )

    if format == "xlsx":
        from io import BytesIO
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Sales by Day"
        ws.append(["date", "revenue"])
        for d in sorted(by_day.keys()):
            ws.append([d, float(by_day[d])])
        ws2 = wb.create_sheet("Item Performance")
        ws2.append(["item", "qty", "revenue"])
        for r in item_perf:
            ws2.append([r["name"], r["qty"], r["revenue"]])
        ws3 = wb.create_sheet("Staff Performance")
        ws3.append(["staff", "tx_count", "revenue"])
        for r in staff_perf:
            ws3.append([r["username"], r["tx_count"], r["revenue"]])
        ws4 = wb.create_sheet("Inventory Valuation")
        ws4.append(["item", "stock_qty", "avg_cost", "value"])
        for r in valuation_rows:
            ws4.append([r["name"], r["stock_qty"], r["avg_cost"], r["value"]])
        ws4.append(["TOTAL", "", "", str(valuation_total)])
        buf = BytesIO()
        wb.save(buf)
        return Response(
            buf.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=geyam_report.xlsx"},
        )

    if format == "pdf":
        from io import BytesIO
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
        buf = BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=A4, title="GEYAM Report")
        styles = getSampleStyleSheet()
        story = [Paragraph("GEYAM Report — Last 30 Days", styles["Title"]), Spacer(1, 12)]

        def _table(title: str, header: list, rows: list):
            story.append(Paragraph(title, styles["Heading2"]))
            data = [header] + rows
            t = Table(data, hAlign="LEFT")
            t.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
            ]))
            story.append(t)
            story.append(Spacer(1, 12))

        _table("Sales by Day", ["Date", "Revenue"],
               [[d, str(by_day[d])] for d in sorted(by_day.keys())])
        _table("Item Performance", ["Item", "Qty", "Revenue"],
               [[r["name"], r["qty"], r["revenue"]] for r in item_perf])
        _table("Staff Performance", ["Staff", "Tx Count", "Revenue"],
               [[r["username"], r["tx_count"], r["revenue"]] for r in staff_perf])
        _table("Inventory Valuation", ["Item", "Stock", "Avg Cost", "Value"],
               [[r["name"], r["stock_qty"], r["avg_cost"], r["value"]] for r in valuation_rows]
               + [["TOTAL", "", "", str(valuation_total)]])
        doc.build(story)
        return Response(
            buf.getvalue(), media_type="application/pdf",
            headers={"Content-Disposition": "attachment; filename=geyam_report.pdf"},
        )

    # CSV (default)
    lines = ["# Sales by Day", "date,revenue"]
    for d in sorted(by_day.keys()):
        lines.append(f"{d},{by_day[d]}")
    lines += ["", "# Item Performance", "item,qty,revenue"]
    for r in item_perf:
        lines.append(f"{r['name']},{r['qty']},{r['revenue']}")
    lines += ["", "# Staff Performance", "staff,tx_count,revenue"]
    for r in staff_perf:
        lines.append(f"{r['username']},{r['tx_count']},{r['revenue']}")
    lines += ["", "# Inventory Valuation", "item,stock_qty,avg_cost,value"]
    for r in valuation_rows:
        lines.append(f"{r['name']},{r['stock_qty']},{r['avg_cost']},{r['value']}")
    lines.append(f"TOTAL,,,{valuation_total}")
    return Response("\n".join(lines), media_type="text/csv",
                    headers={"Content-Disposition": "attachment; filename=geyam_report.csv"})


class AskIn(BaseModel):
    question: str


@router.post("/ask", dependencies=[Depends(require_role("owner"))])
async def ask_endpoint(body: AskIn, tenant_id: int = Depends(get_tenant),
                        session: AsyncSession = Depends(get_session)) -> dict:
    """Owner-side POS analytics assistant, backed by Chenki.

    A keyword classifier maps the question to one of 8 analytics tools
    (product sales, day-by-day revenue, staff, low stock, forecast/reorder,
    detection mix, recent transactions, old transactions); the picked tool
    runs against the tenant-scoped DB and Chenki paraphrases the JSON.
    Tenant scoping comes from the SQLAlchemy event hook via get_tenant.
    """
    return await ask_owner(body.question, session)
